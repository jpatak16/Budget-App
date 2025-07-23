import requests
import os
import time
import datetime
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load variables from .env
load_dotenv()

# Read from environment
access_url = os.getenv("SIMPLEFIN_ACCESS_TOKEN")

# Display error if token not found
if access_url is None:
    raise ValueError("SimpleFIN Access Token was not found in environment.")

# Parse authentication from access URL
scheme, rest = access_url.split("//", 1)
auth, rest = rest.split("@", 1)

username, password = auth.split(":", 1)

url = scheme + "//" + rest + "/accounts"

today = datetime.datetime.now()

# Set start and end dates for our API call
start_dt = datetime.datetime(today.year, today.month-1, 1)
start_unix_ts = int(time.mktime(start_dt.timetuple()))

end_dt = datetime.datetime(today.year, today.month+1, 1)
end_unix_ts = int(time.mktime(end_dt.timetuple()))

# Fetch data from SimpleFIN
response = requests.get(url, auth=(username, password), 
                        params={"start-date": start_unix_ts,
                                "end-date": end_unix_ts}
                        )

# Display error is we fail to feth the accounts
if response.status_code != 200:
    raise Exception(f"Failed to fetch accounts: {response.status_code} - {response.text}")

data = response.json()

# Helper function to convert timestamps into datetime
def ts_to_datetime(ts):
    return datetime.datetime.fromtimestamp(ts)

account_overview = []
transactions = []

# Account name and balance for all accounts
for account in data["accounts"]:
    acct_name = account["name"]
    account_overview.append({
        "acct_name": account["name"],
        "date_updated": ts_to_datetime(account["balance-date"]),
        "balance": account["balance"],
        "date_run": datetime.datetime(today.year, today.month, today.day)
    })
    

    # Info about each transaction within each account
    for trans in account.get("transactions", []):
        transactions.append({
            "account": acct_name,
            "id": trans["id"],
            "description": trans["description"],
            "amount": trans["amount"],
            "payee": trans["payee"],
            "transacted_at": ts_to_datetime(trans["transacted_at"]).date(),
            "category":"",
            "subcategory":""
        })

# Turn our pulled data into pandas dfs
acct_df = pd.DataFrame(account_overview)
trans_df = pd.DataFrame(transactions)

# Import prior data
acct_filename = f"data/account_balances.csv"
trans_log_filename = f"data/trans_log.xlsx"

trans_log_wb = load_workbook(trans_log_filename)
trans_log_ws = trans_log_wb['Sheet1']

old_trans_log_df = pd.read_excel(trans_log_filename)

# Find rows from our new trans file that are not already in the transaction log
new_transactions = trans_df[~trans_df['id'].isin(old_trans_log_df['id'])]

# Merge old trans log and new transactions and sort by date
if len(new_transactions)>0:
    new_trans_log_df = pd.concat([old_trans_log_df, new_transactions])
    new_trans_log_df['transacted_at'] = pd.to_datetime(new_trans_log_df['transacted_at']).dt.date

    # Clear existing highlights from all rows
    no_fill = PatternFill(fill_type=None)
    for row in trans_log_ws.iter_rows():
        for cell in row:
            cell.fill = no_fill

    # Find which rows to highlight
    start_row = trans_log_ws.max_row + 1
    end_row = start_row + len(new_transactions) - 1

    # Replace worksheet with updated data
    trans_log_ws.delete_rows(1, trans_log_ws.max_row)
    trans_log_ws.append(new_trans_log_df.columns.tolist())
    for row in new_trans_log_df.itertuples(index=False):
        trans_log_ws.append(row)

    # Highlight rows that were added the last time the trans log was updated
    new_row_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    for row in trans_log_ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.fill = new_row_fill

# Set column widths for the xlsx file
trans_log_ws.column_dimensions['A'].width = 33
trans_log_ws.column_dimensions['B'].width = 40
trans_log_ws.column_dimensions['C'].width = 69
trans_log_ws.column_dimensions['D'].width = 9
trans_log_ws.column_dimensions['E'].width = 29
trans_log_ws.column_dimensions['F'].width = 12
trans_log_ws.column_dimensions['G'].width = 14
trans_log_ws.column_dimensions['H'].width = 16


# Save
acct_file_exists = os.path.isfile(acct_filename)

acct_df.to_csv(acct_filename, mode='a', index=False, header=not acct_file_exists)
trans_log_wb.save(trans_log_filename)

# Test to see if any accounts on SimpleFIN need attention
status_check = pd.DataFrame(account_overview)

status_check["diff"] = (
    (status_check["date_run"] - status_check["date_updated"])
    .abs()
    .dt.total_seconds() / 86400
)

needs_attention = (status_check["diff"] > 3).sum()

# Confirmation message
print(f" Successfully exported account balances for {len(acct_df)} accounts to {acct_filename}")
print(f" Successfully exported {len(new_transactions)} new transactions to {trans_log_filename}")

if needs_attention > 0:
    print(f"{needs_attention} account(s) need attention (last update > 3 days ago).")
